"""
utils/excel_handler.py — Gestion des fichiers Excel
- Génération de templates de notes pré-remplis (Nom, Prénom, ID)
- Lecture et validation des fichiers importés
- Export de listes d'étudiants
"""

import io
import base64
from datetime import datetime

# openpyxl importé à la demande pour éviter crash si absent
try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


# ─── Couleurs de style Excel ──────────────────────────────────────────────────
COLOR_HEADER_BG  = "1A237E"   # Bleu institutionnel
COLOR_HEADER_FG  = "FFFFFF"
COLOR_SUBHEADER  = "E8EAF6"
COLOR_BORDER     = "BDBDBD"
COLOR_GRADE_COL  = "FFF8E1"   # Fond colonne note à remplir


def _thin_border():
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


# ═══════════════════════════════════════════════════════════════════════════════
# GÉNÉRATION DU TEMPLATE DE NOTES
# ═══════════════════════════════════════════════════════════════════════════════
def generate_grade_template(students, course_code, course_label,
                             eval_label="Évaluation", teacher=""):
    """
    Génère un fichier Excel template pour la saisie des notes.

    Paramètres :
        students      : liste d'objets Student
        course_code   : code du cours (ex: MATH101)
        course_label  : intitulé du cours
        eval_label    : libellé de l'évaluation (ex: Contrôle 1)
        teacher       : nom de l'enseignant

    Retourne :
        bytes : contenu du fichier xlsx en mémoire
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl est requis pour générer des fichiers Excel.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Notes_{course_code}"

    # ── Largeurs des colonnes ─────────────────────────────────────────
    ws.column_dimensions["A"].width = 8    # ID
    ws.column_dimensions["B"].width = 22   # Nom
    ws.column_dimensions["C"].width = 20   # Prénom
    ws.column_dimensions["D"].width = 14   # Note /20
    ws.column_dimensions["E"].width = 14   # Coefficient
    ws.column_dimensions["F"].width = 30   # Commentaire (optionnel)

    # ── En-tête principal ─────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    cell_title = ws["A1"]
    cell_title.value = f"FEUILLE DE NOTES — {course_code} : {course_label}"
    cell_title.font      = Font(bold=True, color=COLOR_HEADER_FG, size=13,
                                name="Calibri")
    cell_title.fill      = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Sous-en-tête infos ────────────────────────────────────────────
    ws.merge_cells("A2:F2")
    info_text = (f"Évaluation : {eval_label}   |   "
                 f"Enseignant : {teacher or '—'}   |   "
                 f"Date de génération : {datetime.now().strftime('%d/%m/%Y')}")
    cell_info = ws["A2"]
    cell_info.value      = info_text
    cell_info.font       = Font(italic=True, size=9, color="555555", name="Calibri")
    cell_info.fill       = PatternFill("solid", fgColor=COLOR_SUBHEADER)
    cell_info.alignment  = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Ligne vide ────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── En-têtes colonnes ─────────────────────────────────────────────
    headers = ["ID", "Nom", "Prénom", "Note /20", "Coefficient", "Commentaire"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font      = Font(bold=True, color=COLOR_HEADER_FG, size=10, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor="283593")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
    ws.row_dimensions[4].height = 22

    # ── Données étudiants ─────────────────────────────────────────────
    for i, student in enumerate(students):
        row = 5 + i
        row_data = [
            student.id,
            student.last_name.upper(),
            student.first_name,
            "",    # Note à remplir
            1.0,   # Coefficient par défaut
            "",    # Commentaire optionnel
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font      = Font(size=10, name="Calibri")
            cell.alignment = Alignment(horizontal="center" if col in (1, 4, 5)
                                       else "left", vertical="center")
            cell.border    = _thin_border()

            # Fond coloré sur la colonne Note
            if col == 4:
                cell.fill = PatternFill("solid", fgColor=COLOR_GRADE_COL)
                # Validation : nombre entre 0 et 20
                from openpyxl.worksheet.datavalidation import DataValidation
                dv = DataValidation(
                    type="decimal",
                    operator="between",
                    formula1=0,
                    formula2=20,
                    showErrorMessage=True,
                    errorTitle="Note invalide",
                    error="La note doit être comprise entre 0 et 20.",
                )
                ws.add_data_validation(dv)
                dv.add(cell)

        ws.row_dimensions[row].height = 20

    # ── Ligne d'instructions ──────────────────────────────────────────
    last_row = 5 + len(students) + 1
    ws.merge_cells(f"A{last_row}:F{last_row}")
    cell_note = ws[f"A{last_row}"]
    cell_note.value = ("INSTRUCTIONS : Remplissez uniquement la colonne "
                       "\"Note /20\" (valeur entre 0 et 20). "
                       "Ne modifiez pas les colonnes ID, Nom, Prénom. "
                       "Enregistrez au format .xlsx avant d'importer.")
    cell_note.font      = Font(italic=True, size=8, color="757575", name="Calibri")
    cell_note.fill      = PatternFill("solid", fgColor="F5F5F5")
    cell_note.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=True)
    ws.row_dimensions[last_row].height = 30

    # ── Figer les volets ──────────────────────────────────────────────
    ws.freeze_panes = "A5"

    # ── Écriture en mémoire ───────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_grade_template_b64(students, course_code, course_label,
                                 eval_label="Évaluation", teacher=""):
    """Retourne le template en base64 pour le téléchargement Dash."""
    content = generate_grade_template(
        students, course_code, course_label, eval_label, teacher
    )
    return base64.b64encode(content).decode()


# ═══════════════════════════════════════════════════════════════════════════════
# LECTURE ET VALIDATION D'UN FICHIER IMPORTÉ
# ═══════════════════════════════════════════════════════════════════════════════
def parse_grade_upload(contents, filename):
    """
    Parse un fichier Excel uploadé via dcc.Upload.

    Paramètres :
        contents : contenu base64 du fichier (fourni par dcc.Upload)
        filename : nom du fichier

    Retourne :
        (True,  list_of_dicts)  en cas de succès
        (False, error_message)  en cas d'erreur

    Chaque dict contient : {student_id, grade, coefficient, comment}
    """
    if not PANDAS_AVAILABLE:
        return False, "pandas est requis pour lire les fichiers Excel."

    # ── Décoder le base64 ──────────────────────────────────────────────
    try:
        content_type, content_string = contents.split(",")
        decoded = base64.b64decode(content_string)
    except Exception:
        return False, "Impossible de décoder le fichier."

    # ── Vérifier l'extension ───────────────────────────────────────────
    if not filename.lower().endswith((".xlsx", ".xls")):
        return False, f"Format non supporté : {filename}. Utilisez un fichier .xlsx."

    # ── Lire avec pandas ──────────────────────────────────────────────
    try:
        df = pd.read_excel(
            io.BytesIO(decoded),
            header=3,          # Les en-têtes sont en ligne 4 (index 3)
            usecols=[0, 1, 2, 3, 4, 5],
        )
    except Exception as e:
        return False, f"Erreur de lecture du fichier Excel : {str(e)}"

    # ── Renommer les colonnes ─────────────────────────────────────────
    try:
        df.columns = ["id", "last_name", "first_name", "grade", "coefficient", "comment"]
    except Exception:
        return False, "Structure du fichier incorrecte. Utilisez le template généré par l'application."

    # ── Nettoyer et valider ───────────────────────────────────────────
    records  = []
    errors   = []
    skipped  = 0

    for idx, row in df.iterrows():
        # Ignorer les lignes vides ou d'instructions
        if pd.isna(row["id"]) or pd.isna(row["grade"]):
            skipped += 1
            continue

        # Valider l'ID étudiant
        try:
            student_id = int(row["id"])
        except (ValueError, TypeError):
            errors.append(f"Ligne {idx + 5} : ID invalide ({row['id']})")
            continue

        # Valider la note
        try:
            grade = float(row["grade"])
            if not (0 <= grade <= 20):
                errors.append(f"Ligne {idx + 5} : Note hors limites ({grade}). Doit être entre 0 et 20.")
                continue
        except (ValueError, TypeError):
            errors.append(f"Ligne {idx + 5} : Note invalide ({row['grade']})")
            continue

        # Coefficient (optionnel)
        try:
            coeff = float(row["coefficient"]) if not pd.isna(row["coefficient"]) else 1.0
            coeff = max(0.1, min(coeff, 10.0))  # Clamp entre 0.1 et 10
        except (ValueError, TypeError):
            coeff = 1.0

        comment = str(row["comment"]).strip() if not pd.isna(row["comment"]) else ""

        records.append({
            "student_id":   student_id,
            "grade":        grade,
            "coefficient":  coeff,
            "comment":      comment,
        })

    if errors:
        error_msg = f"{len(errors)} erreur(s) détectée(s) :\n" + "\n".join(errors[:5])
        if len(errors) > 5:
            error_msg += f"\n... et {len(errors) - 5} autre(s)."
        return False, error_msg

    if not records:
        return False, "Aucune note valide trouvée dans le fichier."

    return True, records


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT LISTE ÉTUDIANTS
# ═══════════════════════════════════════════════════════════════════════════════
def export_students_list(students):
    """
    Exporte la liste complète des étudiants en Excel.
    Retourne les bytes du fichier.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl est requis.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liste Étudiants"

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12

    # Titre
    ws.merge_cells("A1:F1")
    ws["A1"].value      = f"LISTE DES ÉTUDIANTS — Exportée le {datetime.now().strftime('%d/%m/%Y')}"
    ws["A1"].font       = Font(bold=True, color=COLOR_HEADER_FG, size=12, name="Calibri")
    ws["A1"].fill       = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    ws["A1"].alignment  = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # En-têtes
    headers = ["ID", "Nom", "Prénom", "Email", "Date de naissance", "Statut"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = Font(bold=True, color=COLOR_HEADER_FG, size=10, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor="283593")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
    ws.row_dimensions[2].height = 20

    # Données
    for i, s in enumerate(students):
        row = 3 + i
        data = [
            s.id,
            s.last_name.upper(),
            s.first_name,
            s.email,
            str(s.birth_date) if s.birth_date else "",
            "Actif" if s.is_active else "Inactif",
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font      = Font(size=10, name="Calibri")
            cell.alignment = Alignment(horizontal="center" if col in (1, 5, 6)
                                       else "left", vertical="center")
            cell.border    = _thin_border()
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = "A3"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
